import os
import os.path
import openpyxl
import io
import sys
from openpyxl import Workbook
import pandas as pd

book = openpyxl.load_workbook(filename = 'C:/Users/kverner/desktop/ATR_41/material.xlsx')
sheet = book['Sheet2']

pname = ['1A', '1B', '9B', '13A','5I','3I','1I','21I']
for y in pname:
    for k in range(1,42):
        for i in range (14,22):
            B1out = sheet.cell(row=1, column=i)


            B1 = sheet.cell(row=2, column=i)
            mcnp5_file=open('{}ATR.o'.format(k), 'r').readlines()
            
            mcnp5_output_file=os.path.splitext(B1out.value)[0]+'{}_tally.analysis.out'.format(k)

            tally_file=open('C:/Users/kverner/desktop/atr_41/files/{}'.format(mcnp5_output_file),'w+')

            i=0

            for line in mcnp5_file:
                i=i+1
                if  B1.value in line and 'energy' in mcnp5_file[i]:
                    tally_file.write(line)
                    tally_file.write(mcnp5_file[i])
                    for j in range(1,102):
                        tally_file.write(mcnp5_file[i+j])
        
            try:
                try: #this section turns all the output into csv files for calculations
                    fln = "C:/Users/kverner/desktop/atr_41/Files/{}%01d_tally.analysis".format(y) % k
                    df = pd.read_csv(fln+".out", delim_whitespace=True, header=2)
                    df.to_csv(fln+".csv", header=2, index=False)
                except FileNotFoundError:
                    pass
            except pd.io.common.EmptyDataError:
                pass